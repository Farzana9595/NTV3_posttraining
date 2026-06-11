#!/usr/bin/env python3
"""Download reviewed CyVerse/MaizeGDB track candidates from a scanner workbook.

Default behavior is intentionally strict: it only downloads BigWig candidates
with exact SRX or sample-level accession evidence. Use --confidence to include
study_plus_metadata candidates after manual review.
"""

from __future__ import annotations

import argparse
import csv
import os
import sys
import time
import urllib.parse
from pathlib import Path
from typing import Iterable, Optional, Sequence

try:
    import requests
except ImportError as exc:
    raise SystemExit("Missing package: requests. Install with: python -m pip install requests openpyxl") from exc

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("Missing package: openpyxl. Install with: python -m pip install requests openpyxl") from exc


DEFAULT_OUTPUT_DIR = Path(__file__).resolve().parents[1] / "data" / "cyverse_maizegdb_rnaseq_all" / "tracks_raw"
DEFAULT_CONFIDENCE = "exact_srx,exact_sample_accession,sample_name_metadata,study_plus_metadata"
DEFAULT_FILE_TYPES = "BigWig"


def split_csv(value: str) -> set[str]:
    return {part.strip() for part in str(value or "").split(",") if part.strip()}


def read_candidate_rows(workbook: Path, sheet_name: str) -> list[dict[str, object]]:
    wb = load_workbook(workbook, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"Workbook does not contain sheet {sheet_name!r}. Found: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    out = [dict(zip(headers, values)) for values in rows if any(values)]
    wb.close()
    return out


def row_text(row: dict[str, object], key: str) -> str:
    return str(row.get(key) or "").strip()


def as_int(value: object, default: int = 0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def selected_rows(
    rows: Iterable[dict[str, object]],
    file_types: set[str],
    confidence: set[str],
    min_score: int,
) -> list[dict[str, object]]:
    selected: list[dict[str, object]] = []
    seen_urls: set[str] = set()
    for row in rows:
        url = row_text(row, "matched_file_url")
        if not url or url in seen_urls:
            continue
        if file_types and row_text(row, "matched_file_type") not in file_types:
            continue
        if confidence and row_text(row, "match_confidence") not in confidence:
            continue
        if as_int(row.get("score"), 0) < min_score:
            continue
        seen_urls.add(url)
        selected.append(row)
    selected.sort(key=lambda r: (row_text(r, "matched_file_type"), row_text(r, "matched_file_name")))
    return selected


def relative_path_for_url(url: str) -> Path:
    parsed = urllib.parse.urlparse(url)
    path = urllib.parse.unquote(parsed.path)
    marker = "/iplant/home/maizegdb/maizegdb/"
    if marker in path:
        rel = path.split(marker, 1)[1]
    else:
        rel = os.path.basename(path)
    safe_parts = [part for part in rel.split("/") if part and part not in {".", ".."}]
    if not safe_parts:
        safe_parts = [os.path.basename(path) or "downloaded_file"]
    return Path(*safe_parts)


def write_manifest(path: Path, rows: Sequence[dict[str, object]], output_dir: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "file_id",
        "source_excel_row",
        "match_confidence",
        "score",
        "matched_file_type",
        "matched_file_name",
        "matched_file_url",
        "local_path",
        "reason",
    ]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            record = {key: row.get(key, "") for key in fieldnames}
            record["local_path"] = str(output_dir / relative_path_for_url(row_text(row, "matched_file_url")))
            writer.writerow(record)


def download_one(
    session: requests.Session,
    url: str,
    local_path: Path,
    timeout: int,
    overwrite: bool,
) -> tuple[str, int, str]:
    if local_path.exists() and local_path.stat().st_size > 0 and not overwrite:
        return "skipped_existing", local_path.stat().st_size, ""

    local_path.parent.mkdir(parents=True, exist_ok=True)
    part_path = local_path.with_suffix(local_path.suffix + ".part")
    try:
        with session.get(url, stream=True, timeout=timeout) as response:
            response.raise_for_status()
            total = 0
            with part_path.open("wb") as handle:
                for chunk in response.iter_content(chunk_size=1024 * 1024):
                    if not chunk:
                        continue
                    handle.write(chunk)
                    total += len(chunk)
        part_path.replace(local_path)
        return "downloaded", total, ""
    except Exception as exc:  # noqa: BLE001
        return "error", 0, f"{type(exc).__name__}: {exc}"


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Download CyVerse/MaizeGDB track candidates from a scanner workbook.")
    parser.add_argument("--input", required=True, type=Path, help="Scanner workbook with Dataset_Track_Candidates sheet")
    parser.add_argument("--sheet", default="Dataset_Track_Candidates")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--manifest", type=Path, default=None, help="CSV manifest to write before downloading")
    parser.add_argument("--file-types", default=DEFAULT_FILE_TYPES, help="Comma-separated file types to download, e.g. BigWig,BAM")
    parser.add_argument("--confidence", default=DEFAULT_CONFIDENCE, help="Comma-separated confidence labels to allow")
    parser.add_argument("--min-score", type=int, default=70)
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--dry-run", action="store_true", help="Write/list selected files but do not download")
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument("--timeout", type=int, default=300)
    parser.add_argument("--status-csv", type=Path, default=None)
    return parser.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)
    rows = read_candidate_rows(args.input, args.sheet)
    selected = selected_rows(
        rows,
        file_types=split_csv(args.file_types),
        confidence=split_csv(args.confidence),
        min_score=args.min_score,
    )
    if args.limit is not None:
        selected = selected[:args.limit]

    manifest = args.manifest or args.output_dir / "cyverse_track_download_manifest.csv"
    write_manifest(manifest, selected, args.output_dir)
    print(f"Selected {len(selected)} unique candidate URL(s)")
    print(f"Wrote manifest: {manifest}")

    if not selected:
        print("No files matched the download filters. Review confidence/min-score filters or rerun the scanner.")
        return 0
    if args.dry_run:
        for row in selected[:20]:
            print(row_text(row, "matched_file_url"))
        if len(selected) > 20:
            print(f"... {len(selected) - 20} more")
        return 0

    status_path = args.status_csv or args.output_dir / "cyverse_track_download_status.csv"
    status_path.parent.mkdir(parents=True, exist_ok=True)
    session = requests.Session()
    session.headers.update({"User-Agent": "maizegdb-cyverse-track-downloader/1.0"})
    with status_path.open("a", newline="", encoding="utf-8") as handle:
        fieldnames = ["timestamp", "status", "bytes", "local_path", "url", "error"]
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        if handle.tell() == 0:
            writer.writeheader()
        for idx, row in enumerate(selected, start=1):
            url = row_text(row, "matched_file_url")
            local_path = args.output_dir / relative_path_for_url(url)
            print(f"[{idx}/{len(selected)}] {url}")
            status, bytes_written, error = download_one(session, url, local_path, args.timeout, args.overwrite)
            writer.writerow({
                "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
                "status": status,
                "bytes": bytes_written,
                "local_path": str(local_path),
                "url": url,
                "error": error,
            })
            handle.flush()
            if status == "error":
                print(f"ERROR: {error}", file=sys.stderr)
    print(f"Wrote status: {status_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
