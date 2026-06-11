#!/usr/bin/env python3
"""Download ALL BigWig files from CyVerse MaizeGDB.

This script bypasses the scoring/matching system and downloads all BigWig files
found during the scan. Use this when you want the complete CyVerse RNA-seq dataset.

Usage:
    # Download all BigWig files (~1743 files, ~526 GB)
    python download_all_cyverse_bigwig.py --scan-xlsx manifests/cyverse_rnaseq_scan.xlsx

    # Dry run to see what would be downloaded
    python download_all_cyverse_bigwig.py --scan-xlsx manifests/cyverse_rnaseq_scan.xlsx --dry-run

    # Limit to first N files for testing
    python download_all_cyverse_bigwig.py --scan-xlsx manifests/cyverse_rnaseq_scan.xlsx --limit 50
"""

from __future__ import annotations

import argparse
import csv
import os
import sys
import time
import urllib.parse
from pathlib import Path
from typing import Optional, Sequence

try:
    import requests
except ImportError:
    raise SystemExit("Missing package: requests. Install with: pip install requests")

try:
    from openpyxl import load_workbook
except ImportError:
    raise SystemExit("Missing package: openpyxl. Install with: pip install openpyxl")


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_DATA_ROOT = SCRIPT_DIR.parent / "data" / "instadepp_rna_seq"


def read_cyverse_files(xlsx_path: Path, sheet_name: str = "CyVerse_Files") -> list[dict]:
    """Read file records from scanner output. Copies file first to avoid lock issues."""
    import shutil
    import tempfile

    # Copy to temp to avoid PermissionError from OneDrive/Excel lock
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = Path(tmp.name)
    shutil.copy2(xlsx_path, tmp_path)

    try:
        wb = load_workbook(tmp_path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise SystemExit(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        headers = [str(h or "").strip() for h in next(rows)]
        records = []
        for values in rows:
            if any(v is not None for v in values):
                records.append(dict(zip(headers, values)))
        wb.close()
    finally:
        tmp_path.unlink(missing_ok=True)

    return records


RNASEQ_PATH_KEYWORDS = [
    "rnaseq", "rna_seq", "rna-seq", "transcriptome", "expression",
]


def filter_bigwig(files: list[dict], rna_seq_only: bool = True) -> list[dict]:
    """Keep only BigWig files, optionally filtered to RNA-seq paths."""
    bigwig = [f for f in files if str(f.get("file_type", "")).lower() == "bigwig"]
    if not rna_seq_only:
        return bigwig
    return [
        f for f in bigwig
        if any(kw in str(f.get("cyverse_path", "")).lower() for kw in RNASEQ_PATH_KEYWORDS)
    ]


def relative_path_from_cyverse(cyverse_path: str) -> Path:
    """Convert CyVerse path to local relative path."""
    # Remove the /iplant/home/maizegdb/maizegdb/ prefix
    marker = "/iplant/home/maizegdb/maizegdb/"
    if marker in cyverse_path:
        rel = cyverse_path.split(marker, 1)[1]
    else:
        rel = os.path.basename(cyverse_path)

    # Sanitize path parts
    parts = [p for p in rel.split("/") if p and p not in {".", ".."}]
    return Path(*parts) if parts else Path(os.path.basename(cyverse_path))


def download_file(
    session: requests.Session,
    url: str,
    local_path: Path,
    timeout: int = 600,
    overwrite: bool = False,
) -> tuple[str, int, str]:
    """Download a single file from URL."""
    if local_path.exists() and local_path.stat().st_size > 0 and not overwrite:
        return "skipped_existing", local_path.stat().st_size, ""

    local_path.parent.mkdir(parents=True, exist_ok=True)
    part_path = local_path.with_suffix(local_path.suffix + ".part")

    try:
        with session.get(url, stream=True, timeout=timeout) as response:
            response.raise_for_status()
            total = 0
            with part_path.open("wb") as handle:
                for chunk in response.iter_content(chunk_size=1024 * 1024):
                    if chunk:
                        handle.write(chunk)
                        total += len(chunk)
        part_path.replace(local_path)
        return "downloaded", total, ""
    except Exception as e:
        if part_path.exists():
            part_path.unlink()
        return "error", 0, f"{type(e).__name__}: {e}"


def write_manifest(path: Path, files: list[dict], output_dir: Path,
                   pipeline_compat: bool = True) -> None:
    """Write download manifest CSV.

    If pipeline_compat=True, writes in track_download_manifest format so that
    audit_cyverse_bigwig_references.py and subsequent pipeline steps can consume it.
    """
    path.parent.mkdir(parents=True, exist_ok=True)

    if pipeline_compat:
        # Format compatible with audit_cyverse_bigwig_references.py and prepare step
        fieldnames = [
            "file_id", "source_excel_row", "match_confidence", "score",
            "matched_file_type", "matched_file_name", "matched_file_url",
            "local_path", "reason",
        ]
        with path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for i, rec in enumerate(files):
                cyverse_path = str(rec.get("cyverse_path", ""))
                local_rel = relative_path_from_cyverse(cyverse_path)
                writer.writerow({
                    "file_id": rec.get("file_name", ""),
                    "source_excel_row": str(i + 2),
                    "match_confidence": "cyverse_path_rnaseq",
                    "score": "15",
                    "matched_file_type": rec.get("file_type", "BigWig"),
                    "matched_file_name": rec.get("file_name", ""),
                    "matched_file_url": rec.get("file_url", ""),
                    "local_path": str(output_dir / local_rel),
                    "reason": f"CyVerse path: {cyverse_path}",
                })
    else:
        # Raw format
        fieldnames = ["file_url", "cyverse_path", "file_name", "file_type", "size_bytes", "local_path"]
        with path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            for rec in files:
                cyverse_path = str(rec.get("cyverse_path", ""))
                local_rel = relative_path_from_cyverse(cyverse_path)
                writer.writerow({
                    "file_url": rec.get("file_url", ""),
                    "cyverse_path": cyverse_path,
                    "file_name": rec.get("file_name", ""),
                    "file_type": rec.get("file_type", ""),
                    "size_bytes": rec.get("size_bytes", ""),
                    "local_path": str(output_dir / local_rel),
                })


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Download ALL BigWig files from CyVerse MaizeGDB scan results.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    # Download all BigWig files
    python download_all_cyverse_bigwig.py --scan-xlsx manifests/cyverse_rnaseq_scan.xlsx

    # Dry run
    python download_all_cyverse_bigwig.py --scan-xlsx manifests/cyverse_rnaseq_scan.xlsx --dry-run

    # Limit for testing
    python download_all_cyverse_bigwig.py --scan-xlsx manifests/cyverse_rnaseq_scan.xlsx --limit 50
""",
    )
    parser.add_argument("--scan-xlsx", type=Path, required=True,
                        help="Path to cyverse_rnaseq_scan.xlsx from scanner")
    parser.add_argument("--output-dir", type=Path, default=None,
                        help="Output directory (default: data root / tracks_raw)")
    parser.add_argument("--manifest", type=Path, default=None,
                        help="Path to write download manifest CSV (default: manifests/track_download_manifest.csv)")
    parser.add_argument("--status-csv", type=Path, default=None,
                        help="Path to write download status CSV")
    parser.add_argument("--limit", type=int, default=None,
                        help="Limit number of files to download (for testing)")
    parser.add_argument("--all-assays", action="store_true",
                        help="Download BigWig files for ALL assays (not just RNA-seq). Default: RNA-seq only.")
    parser.add_argument("--dry-run", action="store_true",
                        help="Show what would be downloaded without downloading")
    parser.add_argument("--overwrite", action="store_true",
                        help="Overwrite existing files")
    parser.add_argument("--timeout", type=int, default=600,
                        help="Download timeout in seconds (default: 600)")
    return parser.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)
    
    if not args.scan_xlsx.exists():
        print(f"ERROR: Scan file not found: {args.scan_xlsx}", file=sys.stderr)
        return 1
    
    # Set defaults based on scan file location
    data_root = args.scan_xlsx.parent.parent  # manifests -> data root
    output_dir = args.output_dir or data_root / "tracks_raw"
    # Default manifest to pipeline-compatible track_download_manifest.csv
    manifest_path = args.manifest or data_root / "manifests" / "track_download_manifest.csv"
    status_path = args.status_csv or data_root / "manifests" / "all_bigwig_download_status.csv"

    print(f"Reading scan results from: {args.scan_xlsx}")
    all_files = read_cyverse_files(args.scan_xlsx, "CyVerse_Files")
    print(f"Total files in scan: {len(all_files)}")

    rna_seq_only = not args.all_assays
    bigwig_files = filter_bigwig(all_files, rna_seq_only=rna_seq_only)

    if rna_seq_only:
        print(f"RNA-seq BigWig files: {len(bigwig_files)}")
        print(f"  (path contains: {', '.join(RNASEQ_PATH_KEYWORDS)})")
        print(f"  Tip: use --all-assays to include ChIP-seq/ATAC-seq BigWigs too")
    else:
        print(f"All-assay BigWig files: {len(bigwig_files)}")

    if not bigwig_files:
        print("No BigWig files found in scan results!")
        return 1

    # Calculate total size
    total_bytes = sum(int(f.get("size_bytes") or 0) for f in bigwig_files)
    total_gb = total_bytes / (1024**3)
    print(f"Total size: {total_gb:.2f} GB")

    # Apply limit if specified
    if args.limit:
        bigwig_files = bigwig_files[:args.limit]
        print(f"Limited to: {len(bigwig_files)} files")

    # Write manifest in pipeline-compatible format (replaces track_download_manifest.csv)
    write_manifest(manifest_path, bigwig_files, output_dir, pipeline_compat=True)
    print(f"Wrote pipeline manifest: {manifest_path}")

    if args.dry_run:
        print("\n[DRY RUN] Would download:")
        for i, f in enumerate(bigwig_files[:20], 1):
            print(f"  {i}. {f.get('file_name')} ({int(f.get('size_bytes', 0)) / 1e6:.1f} MB)")
        if len(bigwig_files) > 20:
            print(f"  ... and {len(bigwig_files) - 20} more files")
        return 0

    # Download files
    output_dir.mkdir(parents=True, exist_ok=True)
    status_path.parent.mkdir(parents=True, exist_ok=True)

    session = requests.Session()
    session.headers.update({"User-Agent": "maizegdb-bigwig-downloader/1.0"})

    with status_path.open("a", newline="", encoding="utf-8") as handle:
        fieldnames = ["timestamp", "status", "bytes", "local_path", "url", "error"]
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        if handle.tell() == 0:
            writer.writeheader()

        downloaded = 0
        skipped = 0
        errors = 0

        for idx, f in enumerate(bigwig_files, 1):
            url = str(f.get("file_url", ""))
            cyverse_path = str(f.get("cyverse_path", ""))
            local_rel = relative_path_from_cyverse(cyverse_path)
            local_path = output_dir / local_rel

            size_mb = int(f.get("size_bytes", 0)) / 1e6
            print(f"[{idx}/{len(bigwig_files)}] {f.get('file_name')} ({size_mb:.1f} MB)")

            status, bytes_written, error = download_file(
                session, url, local_path, args.timeout, args.overwrite
            )

            writer.writerow({
                "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
                "status": status,
                "bytes": bytes_written,
                "local_path": str(local_path),
                "url": url,
                "error": error,
            })
            handle.flush()

            if status == "downloaded":
                downloaded += 1
                print(f"  Downloaded: {bytes_written / 1e6:.1f} MB")
            elif status == "skipped_existing":
                skipped += 1
                print(f"  Skipped (exists)")
            else:
                errors += 1
                print(f"  ERROR: {error}", file=sys.stderr)

    print(f"\nDone! Downloaded: {downloaded}, Skipped: {skipped}, Errors: {errors}")
    print(f"Status log: {status_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

