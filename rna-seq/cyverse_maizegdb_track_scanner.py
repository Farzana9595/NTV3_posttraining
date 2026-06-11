#!/usr/bin/env python3
"""
CyVerse MaizeGDB Track Scanner

Purpose
-------
Starting from your Excel workbook of maize functional tracks, crawl public
CyVerse/MaizeGDB WebDAV folders and build an Excel report showing which
track/intermediate files appear to be downloadable: BigWig, BAM/CRAM, BED,
bedGraph, GFF/GTF, FASTA, chain, and related JBrowse config files.

This script DOES NOT download large biological data files. It only lists URLs and
creates download commands you can review.

Typical use
-----------
python -m pip install requests openpyxl
python cyverse_maizegdb_track_scanner.py \
  --input "ntv3_maize_posttraining_data_extract(2).xlsx" \
  --output "cyverse_maizegdb_track_search_results.xlsx" \
  --max-depth 3

Deeper NAM/B73v5 scan:
python cyverse_maizegdb_track_scanner.py \
  --input "ntv3_maize_posttraining_data_extract(2).xlsx" \
  --output "cyverse_maizegdb_track_search_results_deep.xlsx" \
  --max-depth 6 \
  --include-folder NAM_PROJECT_JBROWSE_AND_ANALYSES \
  --include-folder B73v5_JBROWSE_AND_ANALYSES
"""

from __future__ import annotations

import argparse
import csv
import html
import os
import re
import sys
import time
import urllib.parse
import xml.etree.ElementTree as ET
from collections import Counter, deque
from dataclasses import dataclass, asdict
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

try:
    import requests
except ImportError as exc:
    raise SystemExit("Missing package: requests. Install with: python -m pip install requests openpyxl") from exc

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:
    raise SystemExit("Missing package: openpyxl. Install with: python -m pip install requests openpyxl") from exc

DEFAULT_ROOT = "https://data.cyverse.org/dav-anon/iplant/home/maizegdb/maizegdb/"
DEFAULT_ALIAS_CSV = os.path.normpath(os.path.join(
    os.path.dirname(__file__),
    "..",
    "maize_rnaseq_pipeline",
    "manifests",
    "maize_rnaseq_ena_aliases.csv",
))
DEFAULT_INCLUDE_FOLDERS = [
    "B73v5_JBROWSE_AND_ANALYSES",
    "NAM_PROJECT_JBROWSE_AND_ANALYSES",
    "OTHER_GENOMES_JBROWSE_AND_ANALYSES",
]
DEFAULT_EXTENSIONS = [
    ".bw", ".bigwig",
    ".bam", ".bai", ".cram", ".crai",
    ".bed", ".bed.gz", ".bedgraph", ".narrowpeak", ".broadpeak",
    ".gff", ".gff3", ".gtf", ".gff3.gz", ".gtf.gz",
    ".fa", ".fasta", ".fna", ".fa.gz", ".fasta.gz", ".fai", ".2bit", ".sizes",
    ".chain", ".chain.gz",
    ".vcf", ".vcf.gz",
    ".json", ".conf", ".txt", ".md",
]

STOPWORDS = {
    "the", "and", "or", "of", "from", "for", "with", "stage", "tissue", "sample",
    "seq", "rna", "chip", "atac", "none", "null", "dap", "days", "after", "pollination",
    "maize", "corn", "zea", "mays", "illumina", "sequencing", "transcriptome",
}

ACCESSION_COLUMNS = [
    "experiment_accession",
    "run_accession",
    "sample_accession",
    "secondary_sample_accession",
    "study_accession",
    "secondary_study_accession",
    "gse",
    "gsm",
]
SAMPLE_LEVEL_PREFIXES = ("SRX", "SRR", "GSM", "SAMN", "SRS", "DRX", "DRR", "ERS", "ERR")
STUDY_LEVEL_PREFIXES = ("GSE", "PRJ", "SRP", "ERP", "DRA", "ERA")
ALIAS_TEXT_COLUMNS = [
    "experiment_alias",
    "sample_alias",
    "experiment_title",
    "study_title",
    "workbook_tissue",
]

@dataclass
class WebDAVEntry:
    url: str
    name: str
    is_dir: bool
    size_bytes: Optional[int] = None
    modified: str = ""

@dataclass
class FileRecord:
    file_url: str
    cyverse_path: str
    file_name: str
    file_type: str
    extension: str
    size_bytes: Optional[int]
    modified: str
    parent_url: str
    source_root: str
    download_command: str

@dataclass
class CandidateRecord:
    source_excel_row: int
    file_id: str
    assay: str
    tissue: str
    experiment_target: str
    dataset: str
    score: int
    match_confidence: str
    matched_search_terms: str
    matched_file_type: str
    matched_file_name: str
    matched_file_url: str
    reason: str
    download_command: str


def ensure_slash(url: str) -> str:
    return url if url.endswith("/") else url + "/"


def normalize_text(value: object) -> str:
    s = str(value or "").lower()
    s = html.unescape(s)
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def tokens(value: object) -> List[str]:
    return [t for t in normalize_text(value).split() if len(t) >= 3 and t not in STOPWORDS]


def compact_token(value: object) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "", str(value or "")).upper()


def unique_strings(values: Iterable[object]) -> List[str]:
    seen = set()
    out: List[str] = []
    for value in values:
        text = str(value or "").strip()
        if not text:
            continue
        if text not in seen:
            seen.add(text)
            out.append(text)
    return out


def detect_extension(name: str) -> str:
    lower = name.lower()
    compound_exts = [
        ".bed.gz", ".gff3.gz", ".gtf.gz", ".fa.gz", ".fasta.gz", ".vcf.gz", ".chain.gz",
        ".narrowpeak", ".broadpeak", ".bedgraph", ".bigwig",
    ]
    for ext in compound_exts:
        if lower.endswith(ext):
            return ext
    root, ext = os.path.splitext(lower)
    return ext


def file_type_from_name(name: str) -> str:
    ext = detect_extension(name)
    lower = name.lower()
    if ext in {".bw", ".bigwig"}:
        return "BigWig"
    if ext == ".bam":
        return "BAM"
    if ext == ".bai":
        return "BAM index"
    if ext == ".cram":
        return "CRAM"
    if ext == ".crai":
        return "CRAM index"
    if ext in {".bed", ".bed.gz", ".narrowpeak", ".broadpeak"}:
        return "BED/Peak"
    if ext == ".bedgraph":
        return "bedGraph"
    if ext in {".gff", ".gff3", ".gff3.gz", ".gtf", ".gtf.gz"}:
        return "GFF/GTF annotation"
    if ext in {".fa", ".fasta", ".fna", ".fa.gz", ".fasta.gz", ".2bit"}:
        return "Genome FASTA/2bit"
    if ext in {".fai", ".sizes"}:
        return "Genome index/chrom sizes"
    if ext in {".chain", ".chain.gz"}:
        return "Liftover chain"
    if ext in {".vcf", ".vcf.gz"}:
        return "VCF"
    if ext in {".json", ".conf"} or "tracklist" in lower:
        return "JBrowse/config"
    if ext in {".txt", ".md"} or "readme" in lower:
        return "README/metadata"
    return "Other"


def path_from_url(url: str) -> str:
    parsed = urllib.parse.urlparse(url)
    # Strip /dav-anon prefix if present so the path looks like iRODS path
    path = urllib.parse.unquote(parsed.path)
    path = re.sub(r"^/dav-anon", "", path)
    return path


def parse_webdav_xml(text: str, base_url: str) -> List[WebDAVEntry]:
    entries: List[WebDAVEntry] = []
    try:
        root = ET.fromstring(text)
    except ET.ParseError:
        return entries

    ns = {"d": "DAV:"}
    base_norm = ensure_slash(base_url)
    base_path = urllib.parse.urlparse(base_norm).path.rstrip("/") + "/"

    for response in root.findall(".//d:response", ns):
        href_el = response.find("d:href", ns)
        if href_el is None or not href_el.text:
            continue
        href = urllib.parse.unquote(href_el.text)
        full_url = urllib.parse.urljoin(base_norm, href)
        # Skip the directory itself
        href_path = urllib.parse.urlparse(full_url).path
        if href_path.rstrip("/") == base_path.rstrip("/"):
            continue

        prop = response.find(".//d:prop", ns)
        name = os.path.basename(href_path.rstrip("/"))
        size_bytes = None
        modified = ""
        is_dir = href_path.endswith("/")
        if prop is not None:
            display = prop.find("d:displayname", ns)
            if display is not None and display.text:
                name = display.text
            res = prop.find("d:resourcetype", ns)
            if res is not None and res.find("d:collection", ns) is not None:
                is_dir = True
            length = prop.find("d:getcontentlength", ns)
            if length is not None and length.text and length.text.isdigit():
                size_bytes = int(length.text)
            lm = prop.find("d:getlastmodified", ns)
            if lm is not None and lm.text:
                modified = lm.text
        if name in {"", ".", ".."}:
            continue
        entries.append(WebDAVEntry(url=full_url, name=name, is_dir=is_dir, size_bytes=size_bytes, modified=modified))
    return entries


def parse_html_listing(text: str, base_url: str) -> List[WebDAVEntry]:
    entries: List[WebDAVEntry] = []
    seen = set()
    # Simple href parser, avoids BeautifulSoup dependency.
    for match in re.finditer(r"href=[\"']([^\"']+)[\"']", text, flags=re.I):
        href = html.unescape(match.group(1))
        if href.startswith("?") or href.startswith("#") or href in {"/", "../", ".."}:
            continue
        full_url = urllib.parse.urljoin(ensure_slash(base_url), href)
        if full_url in seen:
            continue
        seen.add(full_url)
        parsed = urllib.parse.urlparse(full_url)
        if not parsed.scheme.startswith("http"):
            continue
        name = os.path.basename(parsed.path.rstrip("/"))
        if not name or name in {".", ".."}:
            continue
        is_dir = parsed.path.endswith("/") or detect_extension(name) == ""
        entries.append(WebDAVEntry(url=full_url, name=urllib.parse.unquote(name), is_dir=is_dir))
    return entries


def list_webdav_dir(session: requests.Session, url: str, timeout: int = 60) -> Tuple[List[WebDAVEntry], str]:
    """Return entries and method used: PROPFIND, GET, or ERROR."""
    url = ensure_slash(url)
    # PROPFIND is the cleanest way to list WebDAV metadata.
    try:
        r = session.request("PROPFIND", url, headers={"Depth": "1"}, timeout=timeout)
        if r.status_code in {200, 207} and r.text.strip():
            entries = parse_webdav_xml(r.text, url)
            if entries:
                return entries, "PROPFIND"
    except requests.RequestException:
        pass

    # Fallback: CyVerse docs say curl/GET can print HTML directory tables.
    try:
        r = session.get(url, timeout=timeout)
        if r.status_code == 200:
            entries = parse_html_listing(r.text, url)
            return entries, "GET"
        return [], f"HTTP_{r.status_code}"
    except requests.RequestException as exc:
        return [], f"ERROR_{type(exc).__name__}"


def load_dataset_rows(input_xlsx: str, sheet_name: str) -> List[Dict[str, object]]:
    wb = load_workbook(input_xlsx, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"Input workbook does not contain sheet {sheet_name!r}. Found: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    out = []
    for excel_row, values in enumerate(rows_iter, start=2):
        d = dict(zip(headers, values))
        if not any(v is not None and str(v).strip() for v in values):
            continue
        d["source_excel_row"] = excel_row
        out.append(d)
    wb.close()
    return out


def load_alias_rows(alias_csv: str) -> Dict[str, List[Dict[str, str]]]:
    """Load ENA/GEO aliases keyed by SRX experiment accession."""
    aliases: Dict[str, List[Dict[str, str]]] = {}
    if not alias_csv or not os.path.exists(alias_csv):
        return aliases
    with open(alias_csv, newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            experiment = str(row.get("experiment_accession") or "").strip()
            if not experiment:
                continue
            aliases.setdefault(experiment, []).append(row)
    return aliases


def option_values(value: Optional[str]) -> List[str]:
    if value is None:
        return []
    return [v.strip().lower() for v in value.split(",") if v.strip()]


def row_value(row: Dict[str, object], key: str) -> str:
    return str(row.get(key) or "").strip()


def filter_dataset_rows(
    rows: Sequence[Dict[str, object]],
    assay_filter: Optional[str],
    species_filter: Optional[str],
    dataset_filter: Optional[str],
) -> List[Dict[str, object]]:
    assay_values = option_values(assay_filter)
    species_values = option_values(species_filter)
    dataset_values = option_values(dataset_filter)
    out: List[Dict[str, object]] = []
    for row in rows:
        if assay_values and row_value(row, "assay").lower() not in assay_values:
            continue
        if species_values and row_value(row, "specie").lower() not in species_values:
            continue
        if dataset_values and row_value(row, "dataset").lower() not in dataset_values:
            continue
        out.append(row)
    return out


def build_roots(root_url: str, include_folders: Sequence[str]) -> List[str]:
    root = ensure_slash(root_url)
    if include_folders:
        return [urllib.parse.urljoin(root, folder.strip("/") + "/") for folder in include_folders]
    return [root]


def crawl_cyverse(
    root_url: str,
    include_folders: Sequence[str],
    max_depth: int,
    allowed_exts: Sequence[str],
    timeout: int,
    sleep_seconds: float,
    limit_files: Optional[int] = None,
) -> Tuple[List[FileRecord], List[List[object]]]:
    session = requests.Session()
    session.headers.update({"User-Agent": "maizegdb-cyverse-track-scanner/1.0"})
    allowed = {e.lower() for e in allowed_exts}
    roots = build_roots(root_url, include_folders)
    q = deque((ensure_slash(u), 0, ensure_slash(u)) for u in roots)
    visited_dirs = set()
    files: List[FileRecord] = []
    run_log: List[List[object]] = [["timestamp", "event", "url", "depth", "count", "method_or_status"]]

    while q:
        url, depth, source_root = q.popleft()
        if url in visited_dirs:
            continue
        visited_dirs.add(url)
        if depth > max_depth:
            continue
        entries, method = list_webdav_dir(session, url, timeout=timeout)
        run_log.append([time.strftime("%Y-%m-%d %H:%M:%S"), "LIST", url, depth, len(entries), method])
        if sleep_seconds:
            time.sleep(sleep_seconds)

        for entry in entries:
            if entry.is_dir:
                # Avoid accidental traversal outside the root folder.
                if urllib.parse.urlparse(entry.url).netloc == urllib.parse.urlparse(root_url).netloc:
                    if depth < max_depth:
                        q.append((ensure_slash(entry.url), depth + 1, source_root))
                continue
            ext = detect_extension(entry.name)
            if ext.lower() not in allowed:
                continue
            file_type = file_type_from_name(entry.name)
            rec = FileRecord(
                file_url=entry.url,
                cyverse_path=path_from_url(entry.url),
                file_name=entry.name,
                file_type=file_type,
                extension=ext,
                size_bytes=entry.size_bytes,
                modified=entry.modified,
                parent_url=url,
                source_root=source_root,
                download_command=f'curl -L -O "{entry.url}"',
            )
            files.append(rec)
            if limit_files is not None and len(files) >= limit_files:
                run_log.append([time.strftime("%Y-%m-%d %H:%M:%S"), "LIMIT_REACHED", url, depth, len(files), "STOP"])
                return files, run_log
    return files, run_log


def accession_weight(value: str, source_file_id: str) -> int:
    compact = compact_token(value)
    if not compact:
        return 0
    if compact == compact_token(source_file_id):
        return 90
    if compact.startswith(SAMPLE_LEVEL_PREFIXES):
        return 75
    if compact.startswith(STUDY_LEVEL_PREFIXES):
        return 28
    return 0


def search_values_for_row(
    row: Dict[str, object],
    aliases_by_experiment: Dict[str, List[Dict[str, str]]],
) -> Tuple[List[str], List[str]]:
    file_id = row_value(row, "file_id")
    alias_rows = aliases_by_experiment.get(file_id, [])

    accession_values: List[object] = [file_id]
    alias_text_values: List[object] = [
        row.get("tissue"),
        row.get("experiment_target"),
        row.get("download_hint"),
    ]
    for alias_row in alias_rows:
        accession_values.extend(alias_row.get(col, "") for col in ACCESSION_COLUMNS)
        alias_text_values.extend(alias_row.get(col, "") for col in ALIAS_TEXT_COLUMNS)

    accession_terms = unique_strings(accession_values)
    text_terms = unique_strings(alias_text_values)
    return accession_terms, text_terms


def confidence_from_matches(
    matched_accessions: Sequence[str],
    alias_token_matches: int,
    tissue_matches: int,
    source_file_id: str,
) -> str:
    if any(compact_token(v) == compact_token(source_file_id) for v in matched_accessions):
        return "exact_srx"
    if any(compact_token(v).startswith(SAMPLE_LEVEL_PREFIXES) for v in matched_accessions):
        return "exact_sample_accession"
    if any(compact_token(v).startswith(STUDY_LEVEL_PREFIXES) for v in matched_accessions):
        if alias_token_matches or tissue_matches:
            return "study_plus_metadata"
        return "study_only"
    if alias_token_matches >= 2:
        return "sample_name_metadata"
    if tissue_matches:
        return "tissue_only"
    return "weak"


def score_candidate(
    row: Dict[str, object],
    file: FileRecord,
    aliases_by_experiment: Dict[str, List[Dict[str, str]]],
) -> Tuple[int, List[str], List[str], str, bool]:
    path = normalize_text(file.cyverse_path + " " + file.file_name)
    path_compact = compact_token(file.cyverse_path + " " + file.file_name)
    assay = str(row.get("assay") or "")
    target = str(row.get("experiment_target") or "")
    tissue = str(row.get("tissue") or "")
    file_id = str(row.get("file_id") or "")
    score = 0
    reasons: List[str] = []
    matched_terms: List[str] = []
    matched_accessions: List[str] = []
    accession_terms, alias_text_values = search_values_for_row(row, aliases_by_experiment)

    # Direct accession matches are rare in CyVerse, but they are the strongest
    # evidence that a track belongs to a workbook row.
    for term in accession_terms:
        term_compact = compact_token(term)
        weight = accession_weight(term, file_id)
        if weight and term_compact and term_compact in path_compact:
            score += weight
            matched_accessions.append(term)
            matched_terms.append(term)
    if matched_accessions:
        reasons.append("exact accession token(s) in path: " + ", ".join(matched_accessions[:8]))

    # File-type priorities for post-training.
    if file.file_type == "BigWig":
        score += 8
        reasons.append("BigWig direct signal track")
    elif file.file_type in {"BAM", "CRAM"}:
        score += 6
        reasons.append("aligned read file/intermediate")
    elif file.file_type in {"BED/Peak", "bedGraph"}:
        score += 4
        reasons.append("interval/signal intermediate")
    elif file.file_type in {"GFF/GTF annotation", "Genome FASTA/2bit", "Liftover chain"}:
        score += 1
        reasons.append("support/coordinate file")

    # Assay-specific path tokens.
    assay_l = assay.lower()
    if assay_l == "rna-seq":
        if "rna" in path or "rnaseq" in path or "rna seq" in path:
            score += 7
            reasons.append("RNA-seq token in path")
    elif assay_l == "atac-seq":
        if "atac" in path:
            score += 7
            reasons.append("ATAC token in path")
    elif "histone" in assay_l:
        if "chip" in path or "histone" in path:
            score += 5
            reasons.append("histone/ChIP token in path")
        for tok in tokens(target):
            if tok in path:
                score += 5
                reasons.append(f"target token {tok} in path")
    elif "tf" in assay_l:
        if "chip" in path:
            score += 4
            reasons.append("ChIP token in path")
        for tok in tokens(target):
            if tok in path:
                score += 5
                reasons.append(f"TF target token {tok} in path")

    # Tissue tokens. Keep only specific words.
    tissue_matches = 0
    for tok in tokens(tissue):
        if tok in path:
            score += 3
            tissue_matches += 1
            matched_terms.append(tok)
    if tissue_matches:
        reasons.append(f"{tissue_matches} tissue token(s) in path")

    # Sample aliases and titles help when MaizeGDB filenames use lab sample names
    # instead of SRA/GEO accessions. Cap this so broad titles cannot dominate.
    alias_token_hits = 0
    for value in alias_text_values:
        for tok in tokens(value):
            if tok in path:
                alias_token_hits += 1
                matched_terms.append(tok)
                if alias_token_hits <= 6:
                    score += 4
    if alias_token_hits:
        reasons.append(f"{alias_token_hits} alias/title token(s) in path")

    # NAM/B73 folder hints are useful but do not prove sample equivalence.
    if "nam" in path:
        score += 1
        reasons.append("NAM folder/path")
    if "b73" in path:
        score += 1
        reasons.append("B73 folder/path")

    confidence = confidence_from_matches(matched_accessions, alias_token_hits, tissue_matches, file_id)
    return score, reasons, unique_strings(matched_terms), confidence, bool(matched_accessions)


def build_candidates(
    rows: Sequence[Dict[str, object]],
    files: Sequence[FileRecord],
    top_n: int,
    min_score: int,
    aliases_by_experiment: Dict[str, List[Dict[str, str]]],
    require_accession_match: bool,
) -> List[CandidateRecord]:
    candidates: List[CandidateRecord] = []
    for row in rows:
        scored: List[Tuple[int, List[str], List[str], str, FileRecord]] = []
        for f in files:
            score, reasons, matched_terms, confidence, has_accession_match = score_candidate(row, f, aliases_by_experiment)
            if require_accession_match and not has_accession_match:
                continue
            if score >= min_score:
                scored.append((score, reasons, matched_terms, confidence, f))
        scored.sort(key=lambda x: (x[0], x[4].file_type == "BigWig"), reverse=True)
        for score, reasons, matched_terms, confidence, f in scored[:top_n]:
            candidates.append(CandidateRecord(
                source_excel_row=int(row.get("source_excel_row") or 0),
                file_id=str(row.get("file_id") or ""),
                assay=str(row.get("assay") or ""),
                tissue=str(row.get("tissue") or ""),
                experiment_target=str(row.get("experiment_target") or ""),
                dataset=str(row.get("dataset") or ""),
                score=score,
                match_confidence=confidence,
                matched_search_terms="; ".join(matched_terms),
                matched_file_type=f.file_type,
                matched_file_name=f.file_name,
                matched_file_url=f.file_url,
                reason="; ".join(reasons),
                download_command=f'curl -L -O "{f.file_url}"',
            ))
    return candidates


def excel_safe(value: object) -> object:
    if isinstance(value, str) and value.startswith("="):
        return "'" + value
    return value


def write_sheet(ws, headers: Sequence[str], records: Sequence[Dict[str, object]], table_name: str) -> None:
    ws.append(list(headers))
    for rec in records:
        ws.append([excel_safe(rec.get(h, "")) for h in headers])
    style_table(ws, table_name)


def style_table(ws, table_name: str) -> None:
    if ws.max_row >= 1 and ws.max_column >= 1:
        end_cell = f"{get_column_letter(ws.max_column)}{ws.max_row}"
        ref = f"A1:{end_cell}"
        # Excel table names must be alphanumeric/underscore and unique.
        table = Table(displayName=table_name, ref=ref)
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        try:
            ws.add_table(table)
        except ValueError:
            pass
    header_fill = PatternFill("solid", fgColor="115E59")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    # Conservative column widths.
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 10
        for cell in ws[letter][:200]:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(max_len + 2, 60)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def write_output(
    output_xlsx: str,
    dataset_rows: Sequence[Dict[str, object]],
    files: Sequence[FileRecord],
    candidates: Sequence[CandidateRecord],
    run_log: Sequence[Sequence[object]],
    args: argparse.Namespace,
) -> None:
    wb = Workbook()
    # Remove default sheet, then create ordered sheets.
    wb.remove(wb.active)

    summary = wb.create_sheet("Summary")
    file_type_counts = Counter(f.file_type for f in files)
    assay_counts = Counter(str(r.get("assay") or "") for r in dataset_rows)
    summary_rows = [
        ["Metric", "Value"],
        ["Input Excel", args.input],
        ["Input sheet", args.sheet],
        ["Alias CSV", args.alias_csv],
        ["Assay filter", args.assay_filter or ""],
        ["Species filter", args.species_filter or ""],
        ["Dataset filter", args.dataset_filter or ""],
        ["Require accession match", args.require_accession_match],
        ["CyVerse root", args.root_url],
        ["Include folders", "; ".join(args.include_folder or [])],
        ["Max depth", args.max_depth],
        ["Dataset rows", len(dataset_rows)],
        ["CyVerse files listed", len(files)],
        ["Candidate row-file matches", len(candidates)],
        ["BigWig files", file_type_counts.get("BigWig", 0)],
        ["BAM files", file_type_counts.get("BAM", 0)],
        ["CRAM files", file_type_counts.get("CRAM", 0)],
        ["BED/Peak files", file_type_counts.get("BED/Peak", 0)],
        ["GFF/GTF annotation files", file_type_counts.get("GFF/GTF annotation", 0)],
        ["Genome FASTA/2bit files", file_type_counts.get("Genome FASTA/2bit", 0)],
        ["Liftover chain files", file_type_counts.get("Liftover chain", 0)],
        ["RNA-seq rows", assay_counts.get("RNA-seq", 0)],
        ["ATAC-seq rows", assay_counts.get("ATAC-seq", 0)],
        ["Histone ChIP-seq rows", assay_counts.get("Histone ChIP-seq", 0)],
        ["TF ChIP-seq rows", assay_counts.get("TF ChIP-seq", 0)],
        ["Important", "NAM is not a file extension; it is the NAM founder genome/project label."],
        ["Download priority", "BigWig > BAM/CRAM > BED/bedGraph/peaks > FASTQ. GFF/FASTA/chain are support files."],
    ]
    for row in summary_rows:
        summary.append(row)
    style_table(summary, "SummaryTable")

    ds = wb.create_sheet("Dataset_Rows")
    ds_headers = [
        "source_excel_row", "download_hint", "file_id", "biosample_type", "tissue", "assay", "strand",
        "experiment_target", "specie", "dataset", "posttraining_priority", "manual_notes"
    ]
    derived = []
    for r in dataset_rows:
        assay = str(r.get("assay") or "")
        if assay == "RNA-seq":
            priority = "BigWig coverage > BAM/CRAM > FASTQ"
        elif assay == "ATAC-seq":
            priority = "BigWig + BED/peaks > BAM/CRAM > FASTQ"
        elif assay == "Histone ChIP-seq":
            priority = "BigWig + broadPeak/BED > BAM/CRAM > FASTQ"
        elif assay == "TF ChIP-seq":
            priority = "BigWig + narrowPeak/BED > BAM/CRAM > FASTQ"
        else:
            priority = "Tracks/intermediates > FASTQ"
        rr = {h: r.get(h, "") for h in ds_headers}
        rr["posttraining_priority"] = priority
        rr["manual_notes"] = "Excel has SRX/accession; verify genome/cultivar before using a CyVerse NAM/B73 track as equivalent."
        derived.append(rr)
    write_sheet(ds, ds_headers, derived, "DatasetRowsTable")

    fws = wb.create_sheet("CyVerse_Files")
    f_headers = list(asdict(files[0]).keys()) if files else [
        "file_url", "cyverse_path", "file_name", "file_type", "extension", "size_bytes", "modified", "parent_url", "source_root", "download_command"
    ]
    write_sheet(fws, f_headers, [asdict(f) for f in files], "CyVerseFilesTable")

    cws = wb.create_sheet("Dataset_Track_Candidates")
    c_headers = list(asdict(candidates[0]).keys()) if candidates else [
        "source_excel_row", "file_id", "assay", "tissue", "experiment_target", "dataset", "score",
        "match_confidence", "matched_search_terms", "matched_file_type", "matched_file_name",
        "matched_file_url", "reason", "download_command"
    ]
    write_sheet(cws, c_headers, [asdict(c) for c in candidates], "CandidatesTable")

    dws = wb.create_sheet("Download_Commands")
    dws.append(["file_type", "file_name", "file_url", "download_command", "notes"])
    for f in files:
        if f.file_type in {"BigWig", "BAM", "CRAM", "BED/Peak", "bedGraph", "GFF/GTF annotation", "Genome FASTA/2bit", "Liftover chain"}:
            dws.append([f.file_type, f.file_name, f.file_url, f.download_command, "Review coordinate/genome compatibility before using."])
    style_table(dws, "DownloadCommandsTable")

    lws = wb.create_sheet("Run_Log")
    for row in run_log:
        lws.append(list(row))
    style_table(lws, "RunLogTable")

    help_ws = wb.create_sheet("How_To_Read")
    help_rows = [
        ["Topic", "Meaning"],
        ["CyVerse_Files", "Every matching file URL found by crawling the selected CyVerse folders."],
        ["Dataset_Track_Candidates", "Heuristic matches between your Excel rows and CyVerse file paths. These are candidates, not proof of biological equivalence."],
        ["match_confidence", "exact_srx/exact_sample_accession are strongest; study_plus_metadata is usable for review; tissue_only is weak."],
        ["BigWig", "Best direct track label format for post-training."],
        ["BAM/CRAM", "Aligned-read intermediate; can be converted to BigWig."],
        ["GFF3/FASTA/chain", "Support files for coordinates, sequence extraction, or liftover; not RNA-seq signal by themselves."],
        ["NAM", "Nested Association Mapping / founder genomes. It is not a file type."],
    ]
    for row in help_rows:
        help_ws.append(row)
    style_table(help_ws, "HowToReadTable")

    wb.save(output_xlsx)


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Scan CyVerse MaizeGDB WebDAV folders for downloadable track/intermediate files and match them to an Excel dataset.")
    p.add_argument("--input", required=True, help="Original Excel workbook, e.g. ntv3_maize_posttraining_data_extract(2).xlsx")
    p.add_argument("--output", required=True, help="Output Excel tracker with CyVerse file inventory and candidates")
    p.add_argument("--sheet", default="maize_func_tracks", help="Input sheet name")
    p.add_argument("--alias-csv", default=DEFAULT_ALIAS_CSV, help="Optional ENA/GEO alias CSV keyed by SRX experiment_accession")
    p.add_argument("--assay-filter", default=None, help="Comma-separated assay values to keep, e.g. RNA-seq")
    p.add_argument("--species-filter", default=None, help="Comma-separated species values to keep, e.g. zea_mays")
    p.add_argument("--dataset-filter", default=None, help="Comma-separated dataset values to keep, e.g. ncbi_rnaseq")
    p.add_argument("--root-url", default=DEFAULT_ROOT, help="CyVerse WebDAV root URL")
    p.add_argument("--include-folder", action="append", default=None, help="Top-level folder under root to scan. Can be used multiple times. Default scans B73v5, NAM, and OTHER_GENOMES folders.")
    p.add_argument("--max-depth", type=int, default=3, help="Recursive crawl depth under each include folder. Start with 2 or 3; use 6 for deeper scans.")
    p.add_argument("--extensions", default=",".join(DEFAULT_EXTENSIONS), help="Comma-separated extensions to keep")
    p.add_argument("--timeout", type=int, default=60, help="HTTP timeout per directory request")
    p.add_argument("--sleep", type=float, default=0.05, help="Sleep seconds between directory requests")
    p.add_argument("--top-candidates-per-row", type=int, default=5, help="Number of candidate files retained per Excel row")
    p.add_argument("--min-score", type=int, default=8, help="Minimum candidate match score")
    p.add_argument("--require-accession-match", action="store_true", help="Keep only candidates where an SRX/SRR/GSM/SAMN/SRS/GSE/PRJ/SRP token appears in the CyVerse path")
    p.add_argument("--limit-files", type=int, default=None, help="Stop after this many files, useful for testing")
    return p.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)
    if not os.path.exists(args.input):
        print(f"ERROR: input file not found: {args.input}", file=sys.stderr)
        return 2
    include_folders = args.include_folder if args.include_folder is not None else DEFAULT_INCLUDE_FOLDERS
    args.include_folder = include_folders
    allowed_exts = [e.strip().lower() for e in args.extensions.split(",") if e.strip()]

    print(f"Reading dataset rows from {args.input} [{args.sheet}] ...")
    all_dataset_rows = load_dataset_rows(args.input, args.sheet)
    dataset_rows = filter_dataset_rows(
        all_dataset_rows,
        assay_filter=args.assay_filter,
        species_filter=args.species_filter,
        dataset_filter=args.dataset_filter,
    )
    print(f"Loaded {len(dataset_rows)} dataset rows after filters (from {len(all_dataset_rows)} total)")

    print(f"Reading aliases from {args.alias_csv} ...")
    aliases_by_experiment = load_alias_rows(args.alias_csv)
    print(f"Loaded aliases for {len(aliases_by_experiment)} SRX experiments")

    print("Scanning CyVerse folders:")
    for folder in include_folders:
        print(f"  - {urllib.parse.urljoin(ensure_slash(args.root_url), folder.strip('/') + '/')}")
    print(f"Max depth: {args.max_depth}")

    files, run_log = crawl_cyverse(
        root_url=args.root_url,
        include_folders=include_folders,
        max_depth=args.max_depth,
        allowed_exts=allowed_exts,
        timeout=args.timeout,
        sleep_seconds=args.sleep,
        limit_files=args.limit_files,
    )
    print(f"Found {len(files)} file records with requested extensions")
    print("File types:", dict(Counter(f.file_type for f in files)))

    print("Scoring candidate matches to Excel rows ...")
    candidates = build_candidates(
        dataset_rows,
        files,
        top_n=args.top_candidates_per_row,
        min_score=args.min_score,
        aliases_by_experiment=aliases_by_experiment,
        require_accession_match=args.require_accession_match,
    )
    print(f"Candidate matches: {len(candidates)}")

    print(f"Writing {args.output} ...")
    write_output(args.output, dataset_rows, files, candidates, run_log, args)
    print("Done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
